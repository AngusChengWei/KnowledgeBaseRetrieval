"""
文档解析模块 — 支持 PDF、DOCX、TXT、Markdown、Excel 格式的文本提取。

支持的格式:
  - PDF: 使用 PyPDF2 逐页提取文本
  - DOCX/DOC: 使用 python-docx 提取段落文本
  - TXT/MD: 使用 chardet 自动检测编码后读取
  - XLSX/XLS: 使用 openpyxl 逐行提取所有单元格内容

输出格式:
  所有解析函数返回统一的 {"filename": "xxx.pdf", "content": "全文内容"} 结构，
  供 text_splitter 进一步切分。

编码处理:
  TXT 文件编码不确定（UTF-8/GBK/GB2312 等），使用 chardet 自动检测。
"""

import os
from pathlib import Path
from typing import List, Dict

import chardet
import PyPDF2
from docx import Document
import openpyxl


def load_documents(docs_dir: str) -> List[Dict[str, str]]:
    """
    扫描指定目录，解析所有支持的文档。

    参数:
        docs_dir: 文档目录路径

    返回:
        [{"filename": "xxx.pdf", "content": "文档全文内容"}, ...]

    跳过空文件和不支持的格式，解析失败的文件记录日志但不中断整体流程。
    """
    docs_path = Path(docs_dir)
    if not docs_path.exists():
        return []

    documents = []
    supported_extensions = {".pdf", ".docx", ".doc", ".txt", ".md", ".xlsx", ".xls"}

    for file_path in docs_path.iterdir():
        if file_path.is_file() and file_path.suffix.lower() in supported_extensions:
            try:
                content = _parse_file(file_path)
                if content.strip():
                    documents.append({
                        "filename": file_path.name,
                        "content": content
                    })
                    print(f"[文档加载] 成功: {file_path.name} ({len(content)} 字符)")
                else:
                    print(f"[文档加载] 跳过(空内容): {file_path.name}")
            except Exception as e:
                print(f"[文档加载] 失败: {file_path.name}, 错误: {e}")

    print(f"[文档加载] 共加载 {len(documents)} 个文档")
    return documents


def load_single_document(file_path) -> Dict:
    """
    解析单个文件（用于增量更新场景）。

    参数:
        file_path: 文件路径（str 或 Path）

    返回:
        {"filename": "...", "content": "..."} 或 None（文件为空/解析失败）
    """
    fp = Path(file_path)
    try:
        content = _parse_file(fp)
        if content.strip():
            print(f"[文档加载] 单文件成功: {fp.name} ({len(content)} 字符)")
            return {"filename": fp.name, "content": content}
        else:
            print(f"[文档加载] 单文件内容为空: {fp.name}")
    except Exception as e:
        print(f"[文档加载] 单文件解析失败: {fp.name}, {e}")
    return None


def _parse_file(file_path: Path) -> str:
    """根据文件扩展名选择对应的解析器"""
    suffix = file_path.suffix.lower()

    if suffix == ".pdf":
        return _parse_pdf(file_path)
    elif suffix in (".docx", ".doc"):
        return _parse_docx(file_path)
    elif suffix in (".txt", ".md"):
        return _parse_txt(file_path)
    elif suffix in (".xlsx", ".xls"):
        return _parse_excel(file_path)
    else:
        return ""


def _parse_pdf(file_path: Path) -> str:
    """
    解析 PDF 文件 — 使用 PyPDF2 逐页提取文本。

    注意: PyPDF2 对扫描版 PDF（图片型）无效，只能提取文字型 PDF。
    扫描版 PDF 需要 OCR 处理（当前版本不支持）。
    """
    text_parts = []
    with open(file_path, "rb") as f:
        reader = PyPDF2.PdfReader(f)
        for page in reader.pages:
            page_text = page.extract_text()
            if page_text:
                text_parts.append(page_text)
    return "\n".join(text_parts)


def _parse_docx(file_path: Path) -> str:
    """
    解析 DOCX 文件 — 使用 python-docx 提取所有段落文本。

    跳过空段落，段落间用换行符分隔。
    """
    doc = Document(str(file_path))
    paragraphs = [para.text for para in doc.paragraphs if para.text.strip()]
    return "\n".join(paragraphs)


def _parse_txt(file_path: Path) -> str:
    """
    解析 TXT/MD 文件 — chardet 自动检测编码。

    编码检测不是 100% 准确，使用 errors="ignore" 容错处理。
    """
    raw_bytes = file_path.read_bytes()
    detected = chardet.detect(raw_bytes)
    encoding = detected.get("encoding", "utf-8") or "utf-8"
    return raw_bytes.decode(encoding, errors="ignore")


def _parse_excel(file_path: Path) -> str:
    """
    解析 Excel 文件 — 使用 openpyxl 逐行提取所有工作表的单元格内容。

    输出格式:
      [Sheet: 工作表名]
      列1 | 列2 | 列3
      列1 | 列2 | 列3

    read_only=True: 流式读取，适合大文件（不加载整个文件到内存）
    data_only=True: 读取公式的计算结果而非公式本身
    """
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    rows = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows.append(f"[Sheet: {sheet_name}]")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                rows.append(" | ".join(cells))
    wb.close()
    return "\n".join(rows)
